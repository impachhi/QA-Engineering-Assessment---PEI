from __future__ import annotations

import sqlite3
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

from source_data_reader import read_sales_sources


BASE_DIR = Path(__file__).resolve().parent
QA_OUTPUT_DIR = BASE_DIR / "qa_output"

SQLITE_COLUMN_ALIASES = {
    "customers": {
        "customer_id": "customer_id",
        "first": "first_name",
        "last": "last_name",
        "age": "age",
        "country": "country",
    },
    "orders": {
        "order_id": "order_id",
        "item": "item",
        "amount": "amount",
        "customer_id": "customer_id",
    },
    "shipping": {
        "shipping_id": "shipping_id",
        "status": "status",
        "customer_id": "customer_id",
    },
}

DEFAULT_TEST_CASE_WORKBOOK_NAMES = [
    "Test_Cases_ETL_Sales_and_Validation.xlsx",
    "Sales_ETL_Test_Cases_and_Validation.xlsx",
]


def normalize_columns_for_sqlite(df, dataset_name: str):
    """Map source columns to the SQLite names expected by framework SQL."""
    renamed = df.copy()
    renamed.columns = [column.strip().lower() for column in renamed.columns]
    return renamed.rename(columns=SQLITE_COLUMN_ALIASES[dataset_name])


def build_sqlite_connection() -> sqlite3.Connection:
    """Load the known PEI source tables into an in-memory SQLite database."""
    dataframes = read_sales_sources()
    conn = sqlite3.connect(":memory:")
    conn.row_factory = sqlite3.Row

    normalize_columns_for_sqlite(dataframes["customers"], "customers").to_sql(
        "customers", conn, index=False, if_exists="replace"
    )
    normalize_columns_for_sqlite(dataframes["orders"], "orders").to_sql(
        "orders", conn, index=False, if_exists="replace"
    )
    normalize_columns_for_sqlite(dataframes["shipping"], "shipping").to_sql(
        "shipping", conn, index=False, if_exists="replace"
    )
    return conn


def split_sql_statements(sql_text: str) -> list[str]:
    """
    Split SQL text into complete statements using SQLite's parser.

    This is more reliable than assuming each statement ends at a line-ending semicolon.
    """
    statements = []
    buffer = []

    for line in sql_text.splitlines():
        if not line.strip():
            continue
        buffer.append(line)
        candidate = "\n".join(buffer).strip()
        if candidate and sqlite3.complete_statement(candidate):
            statement = candidate.rstrip(";").strip()
            if statement:
                statements.append(statement)
            buffer = []

    trailing = "\n".join(buffer).strip().rstrip(";").strip()
    if trailing:
        statements.append(trailing)

    return statements


def rows_to_markdown(rows: Iterable[sqlite3.Row], headers: list[str]) -> str:
    rows = list(rows)
    if not rows:
        return "_No rows returned._"

    lines = [
        "| " + " | ".join(headers) + " |",
        "| " + " | ".join(["---"] * len(headers)) + " |",
    ]
    for row in rows:
        values = []
        for header in headers:
            value = row[header]
            values.append("NULL" if value is None else str(value).replace("\n", " "))
        lines.append("| " + " | ".join(values) + " |")
    return "\n".join(lines)


def rows_to_console(rows: Iterable[sqlite3.Row], headers: list[str]) -> str:
    rows = list(rows)
    if not rows:
        return "No rows returned."

    widths = [len(header) for header in headers]
    string_rows = []
    for row in rows:
        values = []
        for index, header in enumerate(headers):
            value = "NULL" if row[header] is None else str(row[header])
            widths[index] = max(widths[index], len(value))
            values.append(value)
        string_rows.append(values)

    header_line = " | ".join(header.ljust(widths[index]) for index, header in enumerate(headers))
    separator = "-+-".join("-" * widths[index] for index in range(len(headers)))
    data_lines = [
        " | ".join(value.ljust(widths[index]) for index, value in enumerate(row_values))
        for row_values in string_rows
    ]
    return "\n".join([header_line, separator, *data_lines])


def execute_statement(
    conn: sqlite3.Connection, statement: str
) -> tuple[str, str, list[str], list[sqlite3.Row]]:
    """
    Execute one SQL statement and return console text, markdown text, headers, and rows.
    """
    cursor = conn.execute(statement)
    if cursor.description is None:
        message = "Statement executed successfully with no tabular result."
        return message, f"_{message}_", [], []

    headers = [column[0] for column in cursor.description]
    rows = cursor.fetchall()
    return rows_to_console(rows, headers), rows_to_markdown(rows, headers), headers, rows


def find_first_existing_path(candidates: Iterable[Path]) -> Path | None:
    for candidate in candidates:
        if candidate.exists():
            return candidate
    return None


def resolve_test_case_workbook(
    workbook_path: Path | None = None,
    search_dir: Path = QA_OUTPUT_DIR,
    preferred_names: list[str] | None = None,
) -> Path:
    """
    Resolve the active test case workbook.

    Resolution order:
    1. Explicit path, if provided.
    2. Preferred known workbook names.
    3. Any workbook in qa_output containing a 'Test Cases' sheet.
    """
    if workbook_path:
        if not workbook_path.exists():
            raise FileNotFoundError(f"Workbook not found: {workbook_path}")
        return workbook_path

    preferred = preferred_names or DEFAULT_TEST_CASE_WORKBOOK_NAMES
    candidate_paths = [search_dir / name for name in preferred]
    match = find_first_existing_path(candidate_paths)
    if match:
        return match

    for candidate in sorted(search_dir.glob("*.xlsx")):
        try:
            workbook = load_workbook(candidate, read_only=True)
            if "Test Cases" in workbook.sheetnames:
                return candidate
        except Exception:
            continue

    raise FileNotFoundError(f"No test case workbook with a 'Test Cases' sheet found in {search_dir}")


def read_workbook_rows(workbook_path: Path, sheet_name: str = "Test Cases") -> list[dict[str, str]]:
    workbook = load_workbook(workbook_path)
    worksheet = workbook[sheet_name]

    headers = [worksheet.cell(1, col).value for col in range(1, worksheet.max_column + 1)]
    rows = []
    for row_index in range(2, worksheet.max_row + 1):
        values = [worksheet.cell(row_index, col).value for col in range(1, worksheet.max_column + 1)]
        rows.append(dict(zip(headers, values)))
    return rows
